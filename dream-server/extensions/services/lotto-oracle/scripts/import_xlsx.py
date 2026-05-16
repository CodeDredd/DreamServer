#!/usr/bin/env python3
"""Convert the lotto.de Excel archives into engine-format seed CSVs.

Source files (operator-supplied, dropped into ``seed_data/``):

    LOTTO6aus49_2021.xlsx   one sheet per year (1955–2021)
    Eurojackpot.xlsx        one sheet per year (2012–2021)
    Spiel77.xlsx            one sheet per year (1975–2021)
    SUPER6.xlsx             one sheet per year (1991–2021)

Output (overwrites the small stub seed CSVs already in ``seed_data/``):

    lotto-6aus49.csv  date,Hauptzahlen,Superzahl
    eurojackpot.csv   date,Hauptzahlen,Eurozahlen
    spiel77.csv       date,digits
    super6.csv        date,digits

Per-sheet layout heuristic: walk to col B (the date), then take ints
from col C onwards, skipping ``None`` and string day markers
(``MI``/``SA``/``DI``/``FR``). The number of expected ints is fixed by
the game (6+Zusatz?+Super for lotto, 5+2 for EJ, 7 for spiel77, 6 for
super6). Zusatzzahl (col 7 in lotto pre-2013) is detected by:

    if exactly 8 ints follow → Zusatzzahl present, skip ints[6]
    if exactly 7 ints follow → no Zusatzzahl, ints[6] is Superzahl

After 04.05.2013 the Zusatzzahl was abolished for Lotto 6aus49 (and on
the same date Spiel77/Super6 became Mi+Sa instead of Sa-only).

Run from inside the seed_data dir::

    python3 ../scripts/import_xlsx.py
"""
from __future__ import annotations

import csv
import datetime as dt
import sys
from pathlib import Path

import openpyxl


HERE = Path(__file__).resolve().parent
SEED = HERE.parent / "seed_data"
SOURCES = SEED / "_sources"

DAY_TOKENS = {"MI", "SA", "DI", "FR", "MO", "DO"}


def _ints_after_date(row: tuple) -> list[int]:
    """Return all int cells starting at col C, skipping None and day tokens."""
    out: list[int] = []
    for cell in row[2:]:
        if cell is None:
            continue
        if isinstance(cell, str):
            s = cell.strip().upper()
            if s in DAY_TOKENS or s == "":
                continue
            # money / quote strings (e.g. 'JP', 'unbesetzt')
            break
        if isinstance(cell, float):
            # money columns are floats with cents; counts are ints. Stop at
            # the first non-integer float (e.g. 25_783_219.0 is the
            # Spieleinsatz column).
            if cell.is_integer() and abs(cell) < 100:
                out.append(int(cell))
                continue
            break
        if isinstance(cell, int):
            out.append(cell)
            continue
        break
    return out


def _date_iso(row: tuple) -> str | None:
    cell = row[1]
    if isinstance(cell, dt.datetime):
        return cell.date().isoformat()
    if isinstance(cell, dt.date):
        return cell.isoformat()
    return None


def extract_lotto(wb: openpyxl.Workbook) -> list[tuple[str, list[int], int]]:
    """→ list of (date_iso, sorted_main6, superzahl)."""
    out: list[tuple[str, list[int], int]] = []
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(min_row=5, values_only=True):
            date_iso = _date_iso(row)
            if not date_iso:
                continue
            ints = _ints_after_date(row)
            # Need at least 6 main + 1 super. Accept 7 (no Zusatzzahl) or
            # 8 (with Zusatzzahl, dropped at index 6).
            if len(ints) < 7:
                continue
            main = ints[:6]
            if len(ints) >= 8 and 0 <= ints[7] <= 9 and 1 <= ints[6] <= 49:
                # Zusatzzahl + Superzahl (legacy)
                super_ = ints[7]
            else:
                super_ = ints[6]
            if not (
                len(set(main)) == 6
                and all(1 <= n <= 49 for n in main)
                and 0 <= super_ <= 9
            ):
                continue
            out.append((date_iso, sorted(main), super_))
    return out


def extract_eurojackpot(wb: openpyxl.Workbook) -> list[tuple[str, list[int], list[int]]]:
    out: list[tuple[str, list[int], list[int]]] = []
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(min_row=5, values_only=True):
            date_iso = _date_iso(row)
            if not date_iso:
                continue
            ints = _ints_after_date(row)
            if len(ints) < 7:
                continue
            main = ints[:5]
            euro = ints[5:7]
            # 2012–2014 used 2 aus 8, 2015–2022 2 aus 10, 2022+ 2 aus 12.
            if not (
                len(set(main)) == 5
                and all(1 <= n <= 50 for n in main)
                and len(set(euro)) == 2
                and all(1 <= n <= 12 for n in euro)
            ):
                continue
            out.append((date_iso, sorted(main), sorted(euro)))
    return out


def extract_digit_game(wb: openpyxl.Workbook, digits: int) -> list[tuple[str, str]]:
    out: list[tuple[str, str]] = []
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(min_row=5, values_only=True):
            date_iso = _date_iso(row)
            if not date_iso:
                continue
            ints = _ints_after_date(row)
            if len(ints) < digits:
                continue
            ds = ints[:digits]
            if not all(0 <= d <= 9 for d in ds):
                continue
            out.append((date_iso, "".join(str(d) for d in ds)))
    return out


def _dedup(rows):
    """Keep last occurrence per date (Excel sheets sometimes repeat)."""
    seen = {}
    for r in rows:
        seen[r[0]] = r
    return sorted(seen.values(), key=lambda r: r[0])


def main() -> int:
    files = {
        "lotto":       SOURCES / "LOTTO6aus49_2021.xlsx",
        "eurojackpot": SOURCES / "Eurojackpot.xlsx",
        "spiel77":     SOURCES / "Spiel77.xlsx",
        "super6":      SOURCES / "SUPER6.xlsx",
    }
    for k, fp in files.items():
        if not fp.is_file():
            print(f"  ! missing {fp}", file=sys.stderr)
            return 1

    # ── Lotto 6aus49 ────────────────────────────────────────────────
    wb = openpyxl.load_workbook(files["lotto"], read_only=True, data_only=True)
    rows = _dedup(extract_lotto(wb))
    out = SEED / "lotto-6aus49.csv"
    with out.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["date", "Hauptzahlen", "Superzahl"])
        for date_iso, main, super_ in rows:
            w.writerow([date_iso, " ".join(f"{n:02d}" for n in main), super_])
    print(f"  ✔ lotto-6aus49.csv: {len(rows)} draws "
          f"({rows[0][0]} → {rows[-1][0]})")

    # ── Eurojackpot ────────────────────────────────────────────────
    wb = openpyxl.load_workbook(files["eurojackpot"], read_only=True, data_only=True)
    rows_ej = _dedup(extract_eurojackpot(wb))
    out = SEED / "eurojackpot.csv"
    with out.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["date", "Hauptzahlen", "Eurozahlen"])
        for date_iso, main, euro in rows_ej:
            w.writerow([
                date_iso,
                " ".join(f"{n:02d}" for n in main),
                " ".join(f"{n:02d}" for n in euro),
            ])
    print(f"  ✔ eurojackpot.csv:  {len(rows_ej)} draws "
          f"({rows_ej[0][0]} → {rows_ej[-1][0]})")

    # ── Spiel 77 ───────────────────────────────────────────────────
    wb = openpyxl.load_workbook(files["spiel77"], read_only=True, data_only=True)
    rows_s7 = _dedup(extract_digit_game(wb, 7))
    out = SEED / "spiel77.csv"
    with out.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["date", "digits"])
        for date_iso, digits in rows_s7:
            w.writerow([date_iso, digits])
    print(f"  ✔ spiel77.csv:      {len(rows_s7)} draws "
          f"({rows_s7[0][0]} → {rows_s7[-1][0]})")

    # ── Super 6 ────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(files["super6"], read_only=True, data_only=True)
    rows_s6 = _dedup(extract_digit_game(wb, 6))
    out = SEED / "super6.csv"
    with out.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["date", "digits"])
        for date_iso, digits in rows_s6:
            w.writerow([date_iso, digits])
    print(f"  ✔ super6.csv:       {len(rows_s6)} draws "
          f"({rows_s6[0][0]} → {rows_s6[-1][0]})")

    return 0


if __name__ == "__main__":
    sys.exit(main())

